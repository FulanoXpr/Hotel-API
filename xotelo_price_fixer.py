"""
Hotel Price Fixer - Xotelo API
Searches for hotels that didn't match or didn't get prices one by one.
"""
from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import config
from xotelo_api import XoteloAPI, HotelInfo, RateInfo

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(levelname)s] %(message)s'
)
logger = logging.getLogger(__name__)

# Force immediate output
sys.stdout.reconfigure(line_buffering=True)

# File path defaults
DEFAULT_INPUT_FILE = "PRTC Endorsed Hotels - Updated Xotelo.xlsx"

# Date defaults (relative)
DEFAULT_DAYS_AHEAD = config.DEFAULT_DAYS_AHEAD
DEFAULT_NIGHTS = config.DEFAULT_NIGHTS


def resolve_dates(args: argparse.Namespace) -> Tuple[str, str]:
    """
    Resolve check-in/check-out dates using args or relative defaults.

    Priority:
    1) Explicit --check-in and --check-out
    2) Relative --days-ahead/--nights (or defaults)
    """
    if args.check_in and args.check_out:
        return args.check_in, args.check_out

    today = datetime.now()
    check_in = today + timedelta(days=args.days_ahead)
    check_out = check_in + timedelta(days=args.nights)
    return check_in.strftime("%Y-%m-%d"), check_out.strftime("%Y-%m-%d")


def search_hotel(api: XoteloAPI, query: str) -> Optional[HotelInfo]:
    """
    Search for a hotel by name.

    Args:
        api: XoteloAPI instance
        query: Hotel name to search for

    Returns:
        HotelInfo with key, name, and location, or None if not found
    """
    result = api.search_hotel(query)
    if result:
        return result
    return None


def get_hotel_rates(
    api: XoteloAPI,
    hotel_key: str,
    check_in_date: str,
    check_out_date: str
) -> Optional[RateInfo]:
    """
    Get hotel rates for specific dates.

    Args:
        api: XoteloAPI instance
        hotel_key: Xotelo hotel key

    Returns:
        RateInfo with rate and provider, or None if not available
    """
    return api.get_rates(hotel_key, check_in_date, check_out_date)


def process_unmatched_hotels(
    api: XoteloAPI,
    ws: Worksheet,
    no_match: List[Tuple[int, str]],
    price_col: Optional[int],
    provider_col: Optional[int],
    match_col: Optional[int],
    score_col: Optional[int],
    key_col: Optional[int],
    check_in_date: str,
    check_out_date: str
) -> int:
    """
    Process hotels without a match - search one by one.

    Args:
        api: XoteloAPI instance
        ws: Excel worksheet
        no_match: List of (row, name) tuples for unmatched hotels
        price_col: Column index for price
        provider_col: Column index for provider
        match_col: Column index for match name
        score_col: Column index for match score
        key_col: Column index for hotel key

    Returns:
        Number of hotels updated with prices
    """
    updated = 0

    for idx, (row, name) in enumerate(no_match, 1):
        print(f"\n[{idx}/{len(no_match)}] Searching: {name}")

        # Try different search variations
        search_terms = [
            name,
            name.replace('(', '').replace(')', ''),
            name.split('(')[0].strip(),
            ' '.join(name.split()[:3])  # First 3 words
        ]

        found: Optional[HotelInfo] = None
        for term in search_terms:
            if len(term) < 5:
                continue
            result = search_hotel(api, term)
            if result:
                # Check if it's in Puerto Rico
                location = result.get('location', '').lower()
                if 'puerto rico' in location or 'pr' in location:
                    found = result
                    print(f"  -> Found: {result['name']} ({result['location']})")
                    break
            api.wait()

        if found:
            # Get price
            print("  -> Getting price...")
            rate_data = get_hotel_rates(api, found['key'], check_in_date, check_out_date)

            if rate_data and price_col and provider_col:
                print(f"  -> Price: ${rate_data['rate']} ({rate_data['provider']})")
                ws.cell(row=row, column=price_col, value=rate_data['rate'])
                ws.cell(row=row, column=provider_col, value=rate_data['provider'])
                updated += 1
            elif provider_col:
                print("  -> No price available")
                ws.cell(row=row, column=provider_col, value="No price")

            if match_col:
                ws.cell(row=row, column=match_col, value=found['name'])
            if score_col:
                ws.cell(row=row, column=score_col, value="SEARCH")
            if key_col:
                ws.cell(row=row, column=key_col, value=found['key'])
        else:
            print("  -> NOT FOUND in Puerto Rico")
            if provider_col:
                ws.cell(row=row, column=provider_col, value="Not found")

        api.wait()

    return updated


def process_no_price_hotels(
    api: XoteloAPI,
    ws: Worksheet,
    no_price: List[Tuple[int, str, Optional[str], Optional[str]]],
    price_col: Optional[int],
    provider_col: Optional[int],
    check_in_date: str,
    check_out_date: str
) -> int:
    """
    Process hotels with match but no price - retry getting prices.

    Args:
        api: XoteloAPI instance
        ws: Excel worksheet
        no_price: List of (row, name, match, hotel_key) tuples
        price_col: Column index for price
        provider_col: Column index for provider

    Returns:
        Number of hotels updated with prices
    """
    updated = 0

    for idx, (row, name, match, hotel_key) in enumerate(no_price, 1):
        print(f"\n[{idx}/{len(no_price)}] Retrying: {name}")
        print(f"  -> Key: {hotel_key}")

        if hotel_key:
            rate_data = get_hotel_rates(api, hotel_key, check_in_date, check_out_date)

            if rate_data and price_col and provider_col:
                print(f"  -> Price: ${rate_data['rate']} ({rate_data['provider']})")
                ws.cell(row=row, column=price_col, value=rate_data['rate'])
                ws.cell(row=row, column=provider_col, value=rate_data['provider'])
                updated += 1
            elif provider_col:
                print("  -> Still no price")
                ws.cell(row=row, column=provider_col, value="No price available")
        else:
            print("  -> No hotel key to retry")

        api.wait()

    return updated


def main() -> None:
    """Main execution function."""
    parser = argparse.ArgumentParser(description="Hotel Price Fixer - Xotelo API")
    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT_FILE,
        help="Input Excel file (default: updated Xotelo workbook)"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel file (default: overwrite input)"
    )
    parser.add_argument(
        "--check-in",
        dest="check_in",
        help="Check-in date (YYYY-MM-DD). Requires --check-out if set."
    )
    parser.add_argument(
        "--check-out",
        dest="check_out",
        help="Check-out date (YYYY-MM-DD). Requires --check-in if set."
    )
    parser.add_argument(
        "--days-ahead",
        type=int,
        default=DEFAULT_DAYS_AHEAD,
        help="Days ahead for check-in when dates are not provided."
    )
    parser.add_argument(
        "--nights",
        type=int,
        default=DEFAULT_NIGHTS,
        help="Number of nights when dates are not provided."
    )
    args = parser.parse_args()

    if (args.check_in and not args.check_out) or (args.check_out and not args.check_in):
        raise SystemExit("Both --check-in and --check-out are required together.")

    if args.output is None:
        args.output = args.input

    check_in_date, check_out_date = resolve_dates(args)

    print("=" * 70)
    print("Hotel Price Fixer - Searching Missing Hotels")
    print("=" * 70)
    print(f"Check-in: {check_in_date}")
    print(f"Check-out: {check_out_date}")

    # Initialize API client with longer delay for search operations
    api = XoteloAPI(delay=1.5)

    wb: Workbook = openpyxl.load_workbook(args.input)
    ws: Worksheet = wb.active

    # Find column indices
    headers: Dict[Any, int] = {
        ws.cell(row=1, column=c).value: c
        for c in range(1, ws.max_column + 1)
    }

    name_col = 1
    price_col = headers.get('Xotelo_Price_USD')
    provider_col = headers.get('Provider')
    match_col = headers.get('API_Match_Name')
    score_col = headers.get('Match_Score')
    key_col = headers.get('Hotel_Key')

    print(f"Columns: price={price_col}, match={match_col}, key={key_col}")

    # Collect hotels to process
    no_match: List[Tuple[int, str]] = []
    no_price: List[Tuple[int, str, Optional[str], Optional[str]]] = []

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        if not name:
            continue

        price = ws.cell(row=row, column=price_col).value if price_col else None
        match = ws.cell(row=row, column=match_col).value if match_col else None
        hotel_key = ws.cell(row=row, column=key_col).value if key_col else None

        if not match or match == '':
            no_match.append((row, str(name)))
        elif not price:
            no_price.append((row, str(name), str(match) if match else None, str(hotel_key) if hotel_key else None))

    print(f"\n[INFO] Found {len(no_match)} hotels without match")
    print(f"[INFO] Found {len(no_price)} hotels matched but without price")

    # Process hotels without match - search one by one
    print("\n" + "=" * 70)
    print("PART 1: Searching for unmatched hotels")
    print("=" * 70)

    updated_match = process_unmatched_hotels(
        api,
        ws,
        no_match,
        price_col,
        provider_col,
        match_col,
        score_col,
        key_col,
        check_in_date,
        check_out_date
    )

    # Process hotels with match but no price - retry getting prices
    print("\n" + "=" * 70)
    print("PART 2: Retrying prices for matched hotels")
    print("=" * 70)

    updated_price = process_no_price_hotels(
        api,
        ws,
        no_price,
        price_col,
        provider_col,
        check_in_date,
        check_out_date
    )

    # Save
    wb.save(args.output)
    total_updated = updated_match + updated_price
    print("\n" + "=" * 70)
    print(f"[COMPLETE] Updated {total_updated} hotels with new prices")
    print(f"[SAVED] {args.output}")
    print("=" * 70)


if __name__ == "__main__":
    main()
