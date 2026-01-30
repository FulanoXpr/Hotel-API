"""
Hotel Key Manager - Flask Web UI
Web interface for managing hotel name to Xotelo key mappings.
"""
from __future__ import annotations

import json
import logging
import os
from typing import Any, Dict, List, Optional, Tuple, Union

import requests
from flask import Flask, render_template, request, jsonify
from flask.wrappers import Response

import config

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(levelname)s] %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Configuration from centralized config
EXCEL_FILE = config.EXCEL_FILE
MAPPING_FILE = config.MAPPING_FILE
XOTELO_BASE_URL = config.BASE_URL
API_HOTELS_CACHE = config.API_HOTELS_CACHE


def load_mapping() -> Dict[str, str]:
    """
    Load hotel name to key mapping from JSON file.

    Returns:
        Dictionary mapping hotel names to Xotelo keys
    """
    if os.path.exists(MAPPING_FILE):
        try:
            with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            logger.warning("Could not load mapping: %s", e)
            return {}
    return {}


def save_mapping(mapping: Dict[str, str]) -> None:
    """
    Save hotel name to key mapping to JSON file.

    Args:
        mapping: Dictionary mapping hotel names to Xotelo keys
    """
    with open(MAPPING_FILE, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, indent=4, ensure_ascii=False)


def get_excel_hotels() -> List[str]:
    """
    Get list of hotel names from Excel file.

    Returns:
        Sorted list of unique hotel names
    """
    if not os.path.exists(EXCEL_FILE):
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb.active
        hotels: List[str] = []

        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if name:
                hotels.append(str(name).strip())

        return sorted(list(set(hotels)))
    except (OSError, Exception) as e:
        logger.error("Failed to read Excel file: %s", e)
        return []


@app.route('/')
def index() -> str:
    """Render the main page."""
    return render_template('index.html')


@app.route('/api/hotels')
def get_hotels() -> Response:
    """
    Get list of hotels with their mapping status.

    Returns:
        JSON array of hotel objects with name, key, and status
    """
    excel_hotels = get_excel_hotels()
    mapping = load_mapping()

    result: List[Dict[str, str]] = []
    for name in excel_hotels:
        result.append({
            'name': name,
            'key': mapping.get(name, ""),
            'status': 'mapped' if name in mapping else 'unmapped'
        })

    return jsonify(result)


@app.route('/api/search')
def search_api() -> Union[Tuple[Response, int], Response]:
    """
    Search for hotels in the Xotelo API.

    Query Parameters:
        q: Search query (minimum 2 characters)

    Returns:
        JSON array of matching hotels from API
    """
    query = request.args.get('q', '').lower()
    if not query or len(query) < 2:
        return jsonify([])

    try:
        # Load cached hotels or fetch from API
        all_hotels: List[Dict[str, Any]] = []

        if os.path.exists(API_HOTELS_CACHE):
            with open(API_HOTELS_CACHE, 'r', encoding='utf-8') as f:
                all_hotels = json.load(f)
        else:
            # Fetch hotels from Xotelo /list endpoint (free, no auth needed)
            logger.info("Fetching hotels from Xotelo list API...")
            for offset in range(0, 500, 100):
                response = requests.get(
                    f"{XOTELO_BASE_URL}/list",
                    params={
                        "location_key": config.LOCATION_KEY,
                        "limit": 100,
                        "offset": offset
                    },
                    timeout=15
                )
                response.raise_for_status()
                data = response.json()
                hotels = data.get('result', {}).get('list', [])
                if not hotels:
                    break
                all_hotels.extend(hotels)

            # Cache results
            with open(API_HOTELS_CACHE, 'w', encoding='utf-8') as f:
                json.dump(all_hotels, f, ensure_ascii=False)
            logger.info("Cached %d hotels", len(all_hotels))

        # Filter by query
        results: List[Dict[str, str]] = []
        for hotel in all_hotels:
            name = hotel.get('name', '').lower()
            if query in name:
                results.append({
                    'name': hotel.get('name'),
                    'hotel_key': hotel.get('key'),
                    'short_place_name': hotel.get('location', 'Puerto Rico')
                })

        # Sort by relevance (starts with query first)
        results.sort(key=lambda x: (0 if x['name'].lower().startswith(query) else 1, x['name']))

        return jsonify(results[:20])  # Limit to 20 results

    except requests.exceptions.RequestException as e:
        logger.error("Search request failed: %s", e)
        return jsonify({"error": str(e)}), 500
    except (json.JSONDecodeError, OSError) as e:
        logger.error("Search failed: %s", e)
        return jsonify({"error": str(e)}), 500


@app.route('/api/map', methods=['POST'])
def map_hotel() -> Union[Tuple[Response, int], Response]:
    """
    Map a hotel name to an API key.

    Request Body:
        excel_name: Hotel name from Excel
        api_key: Xotelo hotel key

    Returns:
        JSON success response or error
    """
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    hotel_name = data.get('excel_name')
    hotel_key = data.get('api_key')

    if not hotel_name or not hotel_key:
        return jsonify({"error": "Missing data"}), 400

    mapping = load_mapping()
    mapping[hotel_name] = hotel_key
    save_mapping(mapping)

    return jsonify({"success": True})


@app.route('/api/unmap', methods=['POST'])
def unmap_hotel() -> Response:
    """
    Remove a hotel mapping.

    Request Body:
        excel_name: Hotel name to unmap

    Returns:
        JSON success response
    """
    data = request.json
    if not data:
        return jsonify({"success": True})

    hotel_name = data.get('excel_name')

    mapping = load_mapping()
    if hotel_name in mapping:
        del mapping[hotel_name]
        save_mapping(mapping)

    return jsonify({"success": True})


if __name__ == '__main__':
    print("Starting Hotel Key Manager on http://localhost:5000")
    app.run(debug=False, host='127.0.0.1', port=5000)
