"""
Pestaña de resultados de búsqueda de precios.

Este módulo proporciona la interfaz para visualizar, filtrar y exportar
los resultados de las búsquedas de precios de hoteles.
"""

import os
import sys
from datetime import datetime
from tkinter import filedialog, messagebox
from typing import Any, Callable, Dict, List, Optional

import customtkinter as ctk

sys.path.insert(
    0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
)

from ui.utils.icons import get_icon
from ui.utils.theme import FUENTES, TAMANOS, TemaMode, obtener_tema


class ResultsTab(ctk.CTkFrame):
    """
    Pestaña de resultados de búsqueda.

    Características:
    - Panel resumen con métricas (total, con precio, sin precio)
    - Barras visuales de desglose por proveedor
    - Tabla de resultados ordenable
    - Filtros (con precio / sin precio / todos)
    - Exportar a Excel
    - Copiar al portapapeles
    """

    # Colores por proveedor (cascade + sub-proveedores Xotelo)
    COLORES_PROVEEDORES = {
        # Cascade providers
        "xotelo": "#3498db",
        "serpapi": "#9b59b6",
        "apify": "#e67e22",
        "amadeus": "#1abc9c",
        "cache": "#95a5a6",
        # Sub-proveedores Xotelo (OTAs)
        "vio.com": "#E74C3C",
        "booking.com": "#003580",
        "agoda.com": "#5FC52E",
        "trip.com": "#287DFA",
        "official site": "#F39C12",
        "google hotels": "#4285F4",
        "hotels.com": "#D32F2F",
        "expedia": "#FFCC00",
        "priceline": "#1A4D8F",
        "tripadvisor": "#34E0A1",
        "il hotel": "#FF6B81",
        "destinia": "#FF5722",
        "zenhotels": "#8E44AD",
        "prestigia": "#2ECC71",
        "laterooms": "#E91E63",
        "ostrovok": "#00BCD4",
    }

    def __init__(
        self, master: ctk.CTkFrame, modo_tema: TemaMode = "dark", **kwargs
    ) -> None:
        """
        Inicializa la pestaña de resultados.

        Args:
            master: Widget padre.
            modo_tema: Modo de tema ("dark" o "light").
        """
        self.modo_tema = modo_tema
        self.tema = obtener_tema(modo_tema)

        super().__init__(master, fg_color="transparent", **kwargs)

        # Datos
        self._resultados: List[Dict] = []
        self._resultados_filtrados: List[Dict] = []
        self._filtro_actual = "todos"
        self._orden_columna = "hotel"
        self._orden_ascendente = True

        # Paginación para evitar congelamiento con muchos resultados
        self._filas_por_pagina = 50
        self._filas_mostradas = 0

        # Configurar layout
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # Crear secciones
        self._crear_seccion_resumen()
        self._crear_seccion_filtros()
        self._crear_seccion_tabla()

    def _crear_seccion_resumen(self) -> None:
        """Crea la sección de resumen con métricas."""
        self.frame_resumen = ctk.CTkFrame(
            self,
            fg_color=self.tema["fondo_secundario"],
            corner_radius=TAMANOS["radio_borde"],
        )
        self.frame_resumen.grid(row=0, column=0, sticky="ew", padx=10, pady=10)

        # Título
        ctk.CTkLabel(
            self.frame_resumen,
            text="Results Summary",
            font=FUENTES.get("encabezado", ("Segoe UI", 14, "bold")),
        ).pack(anchor="w", padx=15, pady=(15, 10))

        # Frame de métricas
        self.frame_metricas = ctk.CTkFrame(
            self.frame_resumen,
            fg_color=self.tema["fondo_principal"],
            corner_radius=8,
        )
        self.frame_metricas.pack(fill="x", padx=15, pady=(0, 15))

        for i in range(5):
            self.frame_metricas.grid_columnconfigure(i, weight=1)

        metricas = [
            ("total", "Total", self.tema["texto_principal"]),
            ("con_precio", "With Price", self.tema["estados"]["exito"]),
            ("sin_precio", "No Price", self.tema["estados"]["error"]),
            ("precio_min", "Min Price", self.tema["estados"]["info"]),
            ("precio_max", "Max Price", self.tema["estados"]["warning"]),
        ]

        self._labels_metricas: Dict[str, ctk.CTkLabel] = {}

        for col, (key, label, color) in enumerate(metricas):
            frame = ctk.CTkFrame(self.frame_metricas, fg_color="transparent")
            frame.grid(row=0, column=col, padx=10, pady=10)

            valor_label = ctk.CTkLabel(
                frame,
                text="0" if "precio" not in key else "$0.00",
                font=FUENTES.get("subtitulo", ("Segoe UI", 18, "bold")),
                text_color=color,
            )
            valor_label.pack()

            ctk.CTkLabel(
                frame,
                text=label,
                font=FUENTES.get("pequena", ("Segoe UI", 10)),
            ).pack()

            self._labels_metricas[key] = valor_label

        # Sección de distribución por proveedor
        self._frame_distribucion = ctk.CTkFrame(self.frame_resumen, fg_color="transparent")
        self._frame_distribucion.pack(fill="x", padx=15, pady=(0, 15))

        ctk.CTkLabel(
            self._frame_distribucion,
            text="Distribution by Provider:",
            font=FUENTES.get("normal", ("Segoe UI", 12)),
        ).pack(anchor="w", pady=(0, 5))

        # Barra visual proporcional
        self._frame_barras_proveedores = ctk.CTkFrame(
            self._frame_distribucion,
            fg_color=self.tema["fondo_principal"],
            corner_radius=6,
            height=24,
        )
        self._frame_barras_proveedores.pack(fill="x")
        self._frame_barras_proveedores.pack_propagate(False)

        # Leyenda debajo de la barra
        self._frame_leyenda = ctk.CTkFrame(self._frame_distribucion, fg_color="transparent")
        self._frame_leyenda.pack(fill="x", pady=(6, 0))

    def _crear_seccion_filtros(self) -> None:
        """Crea la sección de filtros y acciones."""
        self.frame_filtros = ctk.CTkFrame(
            self,
            fg_color=self.tema["fondo_secundario"],
            corner_radius=TAMANOS["radio_borde"],
        )
        self.frame_filtros.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))

        frame_inner = ctk.CTkFrame(self.frame_filtros, fg_color="transparent")
        frame_inner.pack(fill="x", padx=15, pady=10)

        # Filtros - segmented button (pills)
        self.var_filtro = ctk.StringVar(value="todos")

        self._FILTRO_MAP = {"All": "todos", "With Price": "con_precio", "No Price": "sin_precio"}

        self.segmented_filtro = ctk.CTkSegmentedButton(
            frame_inner,
            values=["All", "With Price", "No Price"],
            command=self._on_filtro_segmented,
            selected_color=self.tema["acento"],
            selected_hover_color=self.tema["acento_hover"],
            font=FUENTES.get("normal", ("Segoe UI", 12)),
        )
        self.segmented_filtro.set("All")
        self.segmented_filtro.pack(side="left", padx=(0, 10))

        # Botones de acción
        self.boton_copiar = ctk.CTkButton(
            frame_inner,
            text="Copy",
            image=get_icon("copy"),
            compound="left",
            font=FUENTES.get("normal", ("Segoe UI", 12)),
            fg_color=self.tema["acento"],
            hover_color=self.tema["acento_hover"],
            width=100,
            command=self._copiar_portapapeles,
        )
        self.boton_copiar.pack(side="right", padx=(10, 0))

        self.boton_exportar = ctk.CTkButton(
            frame_inner,
            text="Export Excel",
            image=get_icon("export"),
            compound="left",
            font=FUENTES.get("normal", ("Segoe UI", 12)),
            fg_color=self.tema["estados"]["exito"],
            hover_color="#27ae60",
            width=130,
            command=self._exportar_excel,
        )
        self.boton_exportar.pack(side="right")

    def _on_filtro_segmented(self, value: str) -> None:
        """Callback del segmented button de filtro."""
        self.var_filtro.set(self._FILTRO_MAP.get(value, "todos"))
        self._aplicar_filtro()

    def _crear_seccion_tabla(self) -> None:
        """Crea la sección de tabla de resultados."""
        self.frame_tabla = ctk.CTkFrame(
            self,
            fg_color=self.tema["fondo_secundario"],
            corner_radius=TAMANOS["radio_borde"],
        )
        self.frame_tabla.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 10))

        # Encabezado
        self.frame_header = ctk.CTkFrame(
            self.frame_tabla,
            fg_color=self.tema["fondo_principal"],
            corner_radius=0,
        )
        self.frame_header.pack(fill="x", padx=10, pady=(10, 0))

        # Column definitions: (name, weight, anchor)
        # Weights determine proportional width allocation
        self._columnas = [
            ("#", 0.05, "center"),
            ("Hotel", 0.35, "w"),
            ("Price", 0.12, "e"),
            ("Currency", 0.08, "center"),
            ("Provider", 0.13, "center"),
            ("Check-in", 0.13, "center"),
            ("Check-out", 0.13, "center"),
        ]

        # Use grid for proportional columns
        for i, (_, weight, _) in enumerate(self._columnas):
            self.frame_header.grid_columnconfigure(i, weight=int(weight * 100))

        self._header_labels: List[ctk.CTkLabel] = []

        for col_idx, (texto, weight, anchor) in enumerate(self._columnas):
            label = ctk.CTkLabel(
                self.frame_header,
                text=texto,
                font=FUENTES.get("encabezado", ("Segoe UI", 12, "bold")),
                anchor=anchor,
            )
            label.grid(row=0, column=col_idx, sticky="ew", padx=5, pady=8)
            self._header_labels.append(label)

            # Hacer clickeable para ordenar
            if texto not in ["#"]:
                label.bind(
                    "<Button-1>", lambda e, col=texto.lower(): self._ordenar_por(col)
                )
                label.configure(cursor="hand2")

        # Scrollable frame para filas
        self.scroll_frame = ctk.CTkScrollableFrame(
            self.frame_tabla,
            fg_color="transparent",
        )
        self.scroll_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Label para cuando no hay resultados
        self.label_sin_resultados = ctk.CTkLabel(
            self.scroll_frame,
            text="No results to display.\nRun a search in the 'Execute' tab.",
            font=FUENTES.get("normal", ("Segoe UI", 12)),
        )

    def _crear_fila(self, idx: int, resultado: Dict) -> ctk.CTkFrame:
        """Crea una fila de resultado con columnas proporcionales."""
        color_fondo = (
            self.tema["fondo_principal"]
            if idx % 2 == 0
            else self.tema["fondo_secundario"]
        )

        fila = ctk.CTkFrame(
            self.scroll_frame,
            fg_color=color_fondo,
            corner_radius=0,
        )

        # Configure grid columns with same weights as header
        for i, (_, weight, _) in enumerate(self._columnas):
            fila.grid_columnconfigure(i, weight=int(weight * 100))

        font = FUENTES.get("normal", ("Segoe UI", 12))

        # Precio (con manejo de errores para formatos inesperados)
        precio = resultado.get("precio")
        try:
            if precio is not None:
                precio_num = float(precio)
                precio_texto = f"${precio_num:.2f}"
                precio_color = self.tema["estados"]["exito"]
            else:
                precio_texto = "-"
                precio_color = self.tema["texto_secundario"]
        except (ValueError, TypeError):
            precio_texto = str(precio) if precio else "-"
            precio_color = self.tema["texto_secundario"]

        # Proveedor
        proveedor = resultado.get("proveedor") or "-"
        proveedor_color = self.COLORES_PROVEEDORES.get(
            proveedor.lower() if proveedor != "-" else "", self.tema["texto_secundario"]
        )

        # Cell data: (text, color, anchor)
        cells = [
            (str(idx + 1), self.tema["texto_secundario"], "center"),
            (resultado.get("hotel", "")[:50], self.tema["texto_principal"], "w"),
            (precio_texto, precio_color, "e"),
            (resultado.get("moneda") or "-", self.tema["texto_secundario"], "center"),
            (proveedor.capitalize() if proveedor != "-" else "-", proveedor_color, "center"),
            (resultado.get("check_in") or "-", self.tema["texto_secundario"], "center"),
            (resultado.get("check_out") or "-", self.tema["texto_secundario"], "center"),
        ]

        for col_idx, (text, color, anchor) in enumerate(cells):
            ctk.CTkLabel(
                fila,
                text=text,
                font=font,
                text_color=color,
                anchor=anchor,
            ).grid(row=0, column=col_idx, sticky="ew", padx=5, pady=6)

        return fila

    def cargar_resultados(self, resultados: List[Dict]) -> None:
        """
        Carga resultados de búsqueda.

        Args:
            resultados: Lista de diccionarios con resultados.
        """
        # Asegurar que tenemos una lista válida
        self._resultados = resultados if resultados else []
        self._aplicar_filtro()
        self._actualizar_metricas()
        self._actualizar_barras_proveedores()

    def _aplicar_filtro(self) -> None:
        """Aplica el filtro actual y actualiza la tabla."""
        try:
            filtro = self.var_filtro.get()
            self._filtro_actual = filtro

            # Asegurar que _resultados es una lista válida
            if not self._resultados:
                self._resultados_filtrados = []
            elif filtro == "con_precio":
                self._resultados_filtrados = [
                    r for r in self._resultados if r and r.get("precio")
                ]
            elif filtro == "sin_precio":
                self._resultados_filtrados = [
                    r for r in self._resultados if r and not r.get("precio")
                ]
            else:
                self._resultados_filtrados = [r for r in self._resultados if r]

            self._actualizar_tabla()
        except Exception as e:
            # En caso de error, mostrar mensaje y resetear
            print(f"Error al aplicar filtro: {e}")
            self._resultados_filtrados = []
            self._actualizar_tabla()

    def _ordenar_por(self, columna: str) -> None:
        """Ordena los resultados por una columna."""
        if self._orden_columna == columna:
            self._orden_ascendente = not self._orden_ascendente
        else:
            self._orden_columna = columna
            self._orden_ascendente = True

        key_map = {
            "hotel": lambda x: x.get("hotel", "").lower(),
            "precio": lambda x: x.get("precio") or 0,
            "moneda": lambda x: x.get("moneda") or "",
            "proveedor": lambda x: x.get("proveedor") or "",
            "check-in": lambda x: x.get("check_in") or "",
            "check-out": lambda x: x.get("check_out") or "",
        }

        key_func = key_map.get(columna, lambda x: "")
        self._resultados_filtrados.sort(
            key=key_func, reverse=not self._orden_ascendente
        )
        self._actualizar_tabla()

    def _actualizar_tabla(self) -> None:
        """Actualiza la visualización de la tabla con paginación."""
        try:
            # Limpiar filas existentes
            for widget in self.scroll_frame.winfo_children():
                widget.destroy()

            self._filas_mostradas = 0

            if not self._resultados_filtrados:
                self.label_sin_resultados = ctk.CTkLabel(
                    self.scroll_frame,
                    text="No results to display.",
                    font=FUENTES.get("normal", ("Segoe UI", 12)),
                )
                self.label_sin_resultados.pack(pady=50)
                return

            # Mostrar primera página
            self._mostrar_mas_filas()

        except Exception as e:
            print(f"Error al actualizar tabla: {e}")

    def _mostrar_mas_filas(self) -> None:
        """Muestra el siguiente lote de filas."""
        try:
            inicio = self._filas_mostradas
            fin = min(inicio + self._filas_por_pagina, len(self._resultados_filtrados))

            # Crear filas del lote actual
            for idx in range(inicio, fin):
                resultado = self._resultados_filtrados[idx]
                if resultado is None:
                    continue
                try:
                    fila = self._crear_fila(idx, resultado)
                    fila.pack(fill="x", pady=1)
                except Exception as e:
                    print(f"Error al crear fila {idx}: {e}")
                    continue

                # Actualizar UI cada 10 filas para mantener responsividad
                if (idx - inicio) % 10 == 9:
                    self.update_idletasks()

            self._filas_mostradas = fin

            # Si hay más filas, mostrar botón "Cargar más"
            if fin < len(self._resultados_filtrados):
                restantes = len(self._resultados_filtrados) - fin
                self._btn_cargar_mas = ctk.CTkButton(
                    self.scroll_frame,
                    text=f"Load more ({restantes} remaining)",
                    font=FUENTES.get("normal", ("Segoe UI", 12)),
                    fg_color=self.tema["acento"],
                    hover_color=self.tema["acento_hover"],
                    command=self._on_cargar_mas,
                )
                self._btn_cargar_mas.pack(pady=10)

        except Exception as e:
            print(f"Error al mostrar filas: {e}")

    def _on_cargar_mas(self) -> None:
        """Callback para cargar más filas."""
        # Eliminar el botón actual
        if hasattr(self, "_btn_cargar_mas"):
            self._btn_cargar_mas.destroy()
        # Mostrar siguiente lote
        self._mostrar_mas_filas()

    def _actualizar_metricas(self) -> None:
        """Actualiza las métricas del resumen."""
        try:
            if not self._resultados:
                self._labels_metricas["total"].configure(text="0")
                self._labels_metricas["con_precio"].configure(text="0")
                self._labels_metricas["sin_precio"].configure(text="0")
                self._labels_metricas["precio_min"].configure(text="$0.00")
                self._labels_metricas["precio_max"].configure(text="$0.00")
                return

            total = len(self._resultados)
            con_precio = sum(1 for r in self._resultados if r and r.get("precio"))
            sin_precio = total - con_precio

            precios = []
            for r in self._resultados:
                if r and r.get("precio"):
                    try:
                        precios.append(float(r.get("precio")))
                    except (ValueError, TypeError):
                        pass

            precio_min = min(precios) if precios else 0
            precio_max = max(precios) if precios else 0

            self._labels_metricas["total"].configure(text=str(total))
            self._labels_metricas["con_precio"].configure(text=str(con_precio))
            self._labels_metricas["sin_precio"].configure(text=str(sin_precio))
            self._labels_metricas["precio_min"].configure(text=f"${precio_min:.2f}")
            self._labels_metricas["precio_max"].configure(text=f"${precio_max:.2f}")
        except Exception as e:
            print(f"Error al actualizar métricas: {e}")

    def _actualizar_barras_proveedores(self) -> None:
        """Actualiza las barras de distribución por proveedor con leyenda."""
        try:
            # Limpiar barras y leyenda existentes
            for widget in self._frame_barras_proveedores.winfo_children():
                widget.destroy()
            for widget in self._frame_leyenda.winfo_children():
                widget.destroy()

            if not self._resultados:
                return

            # Contar por proveedor
            conteo: Dict[str, int] = {}
            for r in self._resultados:
                if r and r.get("proveedor"):
                    proveedor = str(r.get("proveedor"))
                    conteo[proveedor] = conteo.get(proveedor, 0) + 1

            total_con_precio = sum(conteo.values())
            if total_con_precio == 0:
                return

            # Ordenar por cantidad descendente
            proveedores_ordenados = sorted(conteo.items(), key=lambda x: x[1], reverse=True)

            # Crear barras proporcionales
            relx_actual = 0.0
            for proveedor, cantidad in proveedores_ordenados:
                proporcion = cantidad / total_con_precio
                color = self.COLORES_PROVEEDORES.get(proveedor.lower(), self.tema["acento"])

                barra = ctk.CTkFrame(
                    self._frame_barras_proveedores,
                    fg_color=color,
                    corner_radius=0,
                )
                barra.place(
                    relwidth=max(proporcion, 0.01),  # mínimo visible
                    relheight=1.0,
                    relx=relx_actual,
                )
                relx_actual += proporcion

            # Crear leyenda con dot + nombre + conteo + porcentaje
            for proveedor, cantidad in proveedores_ordenados:
                pct = cantidad / total_con_precio * 100
                color = self.COLORES_PROVEEDORES.get(proveedor.lower(), self.tema["acento"])

                item = ctk.CTkFrame(self._frame_leyenda, fg_color="transparent")
                item.pack(side="left", padx=(0, 16))

                # Dot de color
                dot = ctk.CTkFrame(item, width=10, height=10, corner_radius=5, fg_color=color)
                dot.pack(side="left", padx=(0, 4))
                dot.pack_propagate(False)

                # Texto: "Provider 25 (31.6%)"
                ctk.CTkLabel(
                    item,
                    text=f"{proveedor}  {cantidad} ({pct:.0f}%)",
                    font=FUENTES.get("pequena", ("Segoe UI", 10)),
                ).pack(side="left")

        except Exception as e:
            print(f"Error al actualizar barras de proveedores: {e}")

    def _exportar_excel(self) -> None:
        """Exporta los resultados a un archivo Excel."""
        if not self._resultados:
            messagebox.showinfo("No data", "No results to export.")
            return

        # Pedir ubicación del archivo
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"resultados_hoteles_{fecha}.xlsx",
        )

        if not filename:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados"

            # Encabezados
            headers = [
                "#",
                "Hotel",
                "Price",
                "Currency",
                "Provider",
                "Check-in",
                "Check-out",
            ]
            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Datos
            for idx, resultado in enumerate(self._resultados_filtrados, 2):
                ws.cell(row=idx, column=1, value=idx - 1)
                ws.cell(row=idx, column=2, value=resultado.get("hotel", ""))
                ws.cell(row=idx, column=3, value=resultado.get("precio"))
                ws.cell(row=idx, column=4, value=resultado.get("moneda", ""))
                ws.cell(row=idx, column=5, value=resultado.get("proveedor", ""))
                ws.cell(row=idx, column=6, value=resultado.get("check_in", ""))
                ws.cell(row=idx, column=7, value=resultado.get("check_out", ""))

            # Ajustar anchos
            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 40
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 12
            ws.column_dimensions["G"].width = 12

            wb.save(filename)
            messagebox.showinfo("Exported", f"Results exported to:\n{filename}")

        except ImportError:
            messagebox.showerror(
                "Error", "openpyxl is not installed. Run: pip install openpyxl"
            )
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting: {str(e)}")

    def _copiar_portapapeles(self) -> None:
        """Copia los resultados al portapapeles."""
        if not self._resultados_filtrados:
            messagebox.showinfo("No data", "No results to copy.")
            return

        lineas = ["Hotel\tPrice\tCurrency\tProvider\tCheck-in\tCheck-out"]

        for r in self._resultados_filtrados:
            precio = f"${r.get('precio'):.2f}" if r.get("precio") else "-"
            linea = (
                f"{r.get('hotel', '')}\t{precio}\t{r.get('moneda', '-')}\t"
                f"{r.get('proveedor', '-')}\t{r.get('check_in', '-')}\t{r.get('check_out', '-')}"
            )
            lineas.append(linea)

        texto = "\n".join(lineas)

        self.clipboard_clear()
        self.clipboard_append(texto)

        messagebox.showinfo(
            "Copied",
            f"{len(self._resultados_filtrados)} results copied to clipboard.",
        )

    def cambiar_tema(self, modo_tema: TemaMode) -> None:
        """Cambia el tema de la pestaña."""
        self.modo_tema = modo_tema
        self.tema = obtener_tema(modo_tema)

        # Frames principales
        self.frame_resumen.configure(fg_color=self.tema["fondo_secundario"])
        self.frame_filtros.configure(fg_color=self.tema["fondo_secundario"])
        self.frame_tabla.configure(fg_color=self.tema["fondo_secundario"])

        # Nested frames con colores explícitos
        self.frame_metricas.configure(fg_color=self.tema["fondo_principal"])
        self.frame_header.configure(fg_color=self.tema["fondo_principal"])
        self._frame_barras_proveedores.configure(fg_color=self.tema["fondo_principal"])

        # Actualizar segmented button
        self.segmented_filtro.configure(
            selected_color=self.tema["acento"],
            selected_hover_color=self.tema["acento_hover"],
        )

        # Actualizar colores de métricas (valores con colores semánticos)
        colores_metricas = {
            "total": self.tema["texto_principal"],
            "con_precio": self.tema["estados"]["exito"],
            "sin_precio": self.tema["estados"]["error"],
            "precio_min": self.tema["estados"]["info"],
            "precio_max": self.tema["estados"]["warning"],
        }
        for key, label in self._labels_metricas.items():
            if key in colores_metricas:
                label.configure(text_color=colores_metricas[key])

    def limpiar_resultados(self) -> None:
        """Limpia todos los resultados."""
        self._resultados = []
        self._resultados_filtrados = []
        self._actualizar_tabla()
        self._actualizar_metricas()
