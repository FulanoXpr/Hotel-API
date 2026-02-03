"""
Manejador de archivos Excel para la aplicación Hotel Price Checker.

Este módulo proporciona funcionalidades para cargar, detectar columnas
y guardar datos de hoteles desde/hacia archivos Excel.
"""

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Tipo para representar datos de un hotel
HotelData = Dict[str, Optional[str]]


class ExcelHandler:
    """
    Manejador de archivos Excel para datos de hoteles.

    Proporciona métodos para:
    - Cargar hoteles desde archivos Excel
    - Detectar automáticamente columnas de nombres de hoteles
    - Guardar listas de hoteles a archivos Excel

    Attributes:
        COLUMNAS_NOMBRE_HOTEL: Lista de posibles nombres de columnas para hoteles.
        COLUMNAS_KEY_XOTELO: Lista de posibles nombres de columnas para keys.
        COLUMNAS_BOOKING_URL: Lista de posibles nombres de columnas para URLs.
    """

    # Posibles nombres de columnas para detectar hoteles
    COLUMNAS_NOMBRE_HOTEL: List[str] = [
        "hotel",
        "nombre",
        "name",
        "hotel name",
        "nombre del hotel",
        "property",
        "propiedad",
        "establecimiento",
        "hospederia",
        "hospedería",
        "alojamiento",
        "accommodation",
    ]

    # Posibles nombres de columnas para keys de Xotelo
    COLUMNAS_KEY_XOTELO: List[str] = [
        "key",
        "xotelo_key",
        "xotelo key",
        "hotel_key",
        "hotel key",
        "key_xotelo",
        "clave",
        "clave xotelo",
        "id_xotelo",
        "xotelo_id",
    ]

    # Posibles nombres de columnas para URLs de Booking
    COLUMNAS_BOOKING_URL: List[str] = [
        "url",
        "booking_url",
        "booking url",
        "link",
        "enlace",
        "booking",
        "booking.com",
        "url_booking",
        "website",
    ]

    def __init__(self) -> None:
        """Inicializa el manejador de Excel."""
        pass

    def _normalizar_texto(self, texto: Any) -> str:
        """
        Normaliza un texto para comparación.

        Args:
            texto: Texto a normalizar (puede ser None o cualquier tipo).

        Returns:
            Texto en minúsculas, sin espacios extra, o cadena vacía.
        """
        if texto is None:
            return ""
        return str(texto).lower().strip()

    def _buscar_columna_en_headers(
        self, headers: List[Tuple[int, str]], nombres_posibles: List[str]
    ) -> Optional[int]:
        """
        Busca una columna que coincida con los nombres posibles.

        Args:
            headers: Lista de tuplas (índice, nombre_columna).
            nombres_posibles: Lista de nombres posibles para la columna.

        Returns:
            Índice de la columna encontrada, o None si no se encuentra.
        """
        for idx, header in headers:
            header_normalizado = self._normalizar_texto(header)
            for nombre in nombres_posibles:
                if nombre in header_normalizado or header_normalizado in nombre:
                    return idx
        return None

    def detectar_columna_hotel(self, ruta: str) -> Optional[str]:
        """
        Detecta automáticamente la columna que contiene nombres de hoteles.

        Busca en la primera fila del Excel columnas que coincidan con
        nombres típicos como "Hotel", "Nombre", "Name", etc.

        Args:
            ruta: Ruta al archivo Excel.

        Returns:
            Nombre de la columna detectada, o None si no se encuentra.

        Raises:
            FileNotFoundError: Si el archivo no existe.
            ValueError: Si el archivo no tiene contenido válido.
        """
        path = Path(ruta)
        if not path.exists():
            raise FileNotFoundError(f"El archivo no existe: {ruta}")

        workbook = load_workbook(ruta, read_only=True, data_only=True)
        sheet: Worksheet = workbook.active

        if sheet is None:
            workbook.close()
            raise ValueError("El archivo Excel no tiene hojas activas.")

        # Obtener la primera fila (headers)
        primera_fila = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        workbook.close()

        if not primera_fila or not primera_fila[0]:
            raise ValueError("El archivo Excel no tiene encabezados.")

        headers = primera_fila[0]

        # Buscar columna que coincida
        for idx, header in enumerate(headers):
            if header is None:
                continue
            header_normalizado = self._normalizar_texto(header)
            for nombre_posible in self.COLUMNAS_NOMBRE_HOTEL:
                if nombre_posible in header_normalizado:
                    return str(header)

        return None

    def _detectar_columnas(
        self, headers: Tuple[Any, ...]
    ) -> Tuple[Optional[int], Optional[int], Optional[int]]:
        """
        Detecta los índices de las columnas relevantes.

        Args:
            headers: Tupla con los encabezados de las columnas.

        Returns:
            Tupla con (índice_nombre, índice_key, índice_url).
        """
        # Crear lista de (índice, nombre) para headers válidos
        headers_list: List[Tuple[int, str]] = [
            (idx, str(h)) for idx, h in enumerate(headers) if h is not None
        ]

        idx_nombre = self._buscar_columna_en_headers(
            headers_list, self.COLUMNAS_NOMBRE_HOTEL
        )
        idx_key = self._buscar_columna_en_headers(
            headers_list, self.COLUMNAS_KEY_XOTELO
        )
        idx_url = self._buscar_columna_en_headers(
            headers_list, self.COLUMNAS_BOOKING_URL
        )

        return idx_nombre, idx_key, idx_url

    def cargar_excel(self, ruta: str) -> List[HotelData]:
        """
        Carga la lista de hoteles desde un archivo Excel.

        Detecta automáticamente las columnas de nombre, key Xotelo y URL.

        Args:
            ruta: Ruta al archivo Excel.

        Returns:
            Lista de diccionarios con keys: nombre, xotelo_key, booking_url.

        Raises:
            FileNotFoundError: Si el archivo no existe.
            ValueError: Si no se puede detectar la columna de hoteles.
        """
        path = Path(ruta)
        if not path.exists():
            raise FileNotFoundError(f"El archivo no existe: {ruta}")

        workbook = load_workbook(ruta, read_only=True, data_only=True)
        sheet: Worksheet = workbook.active

        if sheet is None:
            workbook.close()
            raise ValueError("El archivo Excel no tiene hojas activas.")

        filas = list(sheet.iter_rows(values_only=True))
        workbook.close()

        if not filas:
            raise ValueError("El archivo Excel está vacío.")

        # Detectar columnas
        headers = filas[0]
        idx_nombre, idx_key, idx_url = self._detectar_columnas(headers)

        if idx_nombre is None:
            raise ValueError(
                "No se pudo detectar la columna de nombres de hoteles. "
                "Asegúrese de que exista una columna con nombre: "
                f"{', '.join(self.COLUMNAS_NOMBRE_HOTEL[:5])}"
            )

        hoteles: List[HotelData] = []

        # Procesar filas de datos (saltando header)
        for fila in filas[1:]:
            if not fila or len(fila) <= idx_nombre:
                continue

            nombre = fila[idx_nombre]
            if not nombre or self._normalizar_texto(nombre) == "":
                continue

            # Obtener key si existe la columna
            xotelo_key: Optional[str] = None
            if idx_key is not None and len(fila) > idx_key:
                valor_key = fila[idx_key]
                if valor_key is not None and str(valor_key).strip():
                    xotelo_key = str(valor_key).strip()

            # Obtener URL si existe la columna
            booking_url: Optional[str] = None
            if idx_url is not None and len(fila) > idx_url:
                valor_url = fila[idx_url]
                if valor_url is not None and str(valor_url).strip():
                    booking_url = str(valor_url).strip()

            hoteles.append(
                {
                    "nombre": str(nombre).strip(),
                    "xotelo_key": xotelo_key,
                    "booking_url": booking_url,
                }
            )

        return hoteles

    def guardar_excel(self, ruta: str, hoteles: List[HotelData]) -> None:
        """
        Guarda la lista de hoteles a un archivo Excel.

        Crea un archivo con columnas: Hotel, Xotelo Key, Booking URL.

        Args:
            ruta: Ruta donde guardar el archivo Excel.
            hoteles: Lista de diccionarios con datos de hoteles.

        Raises:
            ValueError: Si la lista de hoteles está vacía.
            PermissionError: Si no se puede escribir en la ruta especificada.
        """
        if not hoteles:
            raise ValueError("La lista de hoteles está vacía.")

        workbook = Workbook()
        sheet: Worksheet = workbook.active
        sheet.title = "Hoteles"

        # Escribir encabezados
        headers = ["Hotel", "Xotelo Key", "Booking URL"]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Escribir datos
        for row_idx, hotel in enumerate(hoteles, start=2):
            sheet.cell(row=row_idx, column=1, value=hotel.get("nombre", ""))
            sheet.cell(row=row_idx, column=2, value=hotel.get("xotelo_key", ""))
            sheet.cell(row=row_idx, column=3, value=hotel.get("booking_url", ""))

        # Ajustar ancho de columnas
        sheet.column_dimensions["A"].width = 40
        sheet.column_dimensions["B"].width = 25
        sheet.column_dimensions["C"].width = 50

        workbook.save(ruta)
        workbook.close()

    def exportar_seleccionados(
        self, ruta: str, hoteles: List[HotelData], solo_con_key: bool = False
    ) -> int:
        """
        Exporta hoteles seleccionados a un archivo Excel.

        Args:
            ruta: Ruta donde guardar el archivo.
            hoteles: Lista de hoteles a exportar.
            solo_con_key: Si True, solo exporta hoteles con key Xotelo.

        Returns:
            Número de hoteles exportados.
        """
        if solo_con_key:
            hoteles_filtrados = [h for h in hoteles if h.get("xotelo_key")]
        else:
            hoteles_filtrados = hoteles

        if hoteles_filtrados:
            self.guardar_excel(ruta, hoteles_filtrados)

        return len(hoteles_filtrados)


# Instancia global para uso conveniente
excel_handler = ExcelHandler()
