"""
Hotel Price Updater - Xotelo API (Key-Based)
Fetches hotel prices from Xotelo using direct hotel keys from hotel_keys_db.json.
Supports both interactive and automatic modes for monthly automation.

Usage:
  Interactive:  python xotelo_price_updater.py
  Automatic:    python xotelo_price_updater.py --auto
  Multi-date:   python xotelo_price_updater.py --auto --multi-date
  Cascade:      python xotelo_price_updater.py --auto --cascade
  Full:         python xotelo_price_updater.py --auto --multi-date --cascade
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple, TypedDict

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import config
from xotelo_api import XoteloAPI, RateInfo, get_client

# Import cascade pipeline components (optional)
try:
    from price_providers import (
        CascadePriceProvider,
        PriceCache,
        XoteloProvider,
        SerpApiProvider,
        ApifyProvider,
        PriceResult
    )
    CASCADE_AVAILABLE = True
except ImportError:
    CASCADE_AVAILABLE = False

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(levelname)s] %(message)s'
)
logger = logging.getLogger(__name__)

# Force immediate output
sys.stdout.reconfigure(line_buffering=True)

# Re-export config values for backwards compatibility
BASE_URL = config.BASE_URL
EXCEL_FILE = config.EXCEL_FILE
HOTEL_KEYS_DB = config.HOTEL_KEYS_DB
DEFAULT_DAYS_AHEAD = config.DEFAULT_DAYS_AHEAD
DEFAULT_NIGHTS = config.DEFAULT_NIGHTS
DEFAULT_ROOMS = config.DEFAULT_ROOMS
DEFAULT_ADULTS = config.DEFAULT_ADULTS
TIMEOUT = config.TIMEOUT
REQUEST_DELAY = config.REQUEST_DELAY


class SearchParams(TypedDict):
    """Type definition for search parameters."""
    chk_in: str
    chk_out: str
    rooms: int
    adults: int
    nights: int


class HotelPriceData(TypedDict, total=False):
    """Type definition for hotel price data."""
    price: Optional[float]
    provider: str
    hotel_key: str
    date_used: str  # Which date range found the price
    source: str  # Which cascade provider found the price


class DateRange(TypedDict):
    """Type definition for a date range to try."""
    label: str
    chk_in: str
    chk_out: str


def get_user_input() -> Optional[SearchParams]:
    """Get search parameters from user (interactive mode)."""
    print("\n" + "=" * 70)
    print("CONFIGURACION DE BUSQUEDA")
    print("=" * 70)

    # Get check-in date
    while True:
        default_date = (datetime.now() + timedelta(days=DEFAULT_DAYS_AHEAD)).strftime("%Y-%m-%d")
        print(f"\nFecha de Check-in (YYYY-MM-DD) [default: {default_date}]: ", end="")
        chk_in = input().strip()
        if not chk_in:
            chk_in = default_date

        try:
            check_in_date = datetime.strptime(chk_in, "%Y-%m-%d")
            if check_in_date < datetime.now():
                print("   La fecha debe ser en el futuro. Intenta de nuevo.")
                continue
            break
        except ValueError:
            print("   Formato invalido. Usa YYYY-MM-DD (ej: 2026-02-15)")

    # Get check-out date
    while True:
        default_checkout = (check_in_date + timedelta(days=DEFAULT_NIGHTS)).strftime("%Y-%m-%d")
        print(f"\nFecha de Check-out (YYYY-MM-DD) [default: {default_checkout}]: ", end="")
        chk_out = input().strip()
        if not chk_out:
            chk_out = default_checkout

        try:
            check_out_date = datetime.strptime(chk_out, "%Y-%m-%d")
            if check_out_date <= check_in_date:
                print("   Check-out debe ser despues de check-in. Intenta de nuevo.")
                continue
            break
        except ValueError:
            print("   Formato invalido. Usa YYYY-MM-DD (ej: 2026-02-16)")

    # Get number of rooms
    while True:
        print(f"\nNumero de habitaciones (1-8) [default: {DEFAULT_ROOMS}]: ", end="")
        rooms_input = input().strip()
        if not rooms_input:
            rooms = DEFAULT_ROOMS
            break
        try:
            rooms = int(rooms_input)
            if 1 <= rooms <= 8:
                break
            print("   Debe ser entre 1 y 8 habitaciones.")
        except ValueError:
            print("   Por favor ingresa un numero valido.")

    # Get number of adults per room
    while True:
        print(f"\nAdultos por habitacion (1-4) [default: {DEFAULT_ADULTS}]: ", end="")
        adults_input = input().strip()
        if not adults_input:
            adults = DEFAULT_ADULTS
            break
        try:
            adults = int(adults_input)
            if 1 <= adults <= 4:
                break
            print("   Debe ser entre 1 y 4 adultos por habitacion.")
        except ValueError:
            print("   Por favor ingresa un numero valido.")

    nights = (check_out_date - check_in_date).days

    print("\n" + "-" * 70)
    print("RESUMEN DE BUSQUEDA:")
    print(f"   * Check-in:    {chk_in}")
    print(f"   * Check-out:   {chk_out}")
    print(f"   * Noches:      {nights}")
    print(f"   * Habitaciones: {rooms}")
    print(f"   * Adultos/hab: {adults}")
    print("-" * 70)

    print("\nContinuar con esta busqueda? (S/n): ", end="")
    confirm = input().strip().lower()
    if confirm == 'n':
        print("\nBusqueda cancelada.")
        return None

    return SearchParams(
        chk_in=chk_in,
        chk_out=chk_out,
        rooms=rooms,
        adults=adults,
        nights=nights
    )


def get_auto_params() -> SearchParams:
    """Get default parameters for automatic mode."""
    today = datetime.now()
    check_in = today + timedelta(days=DEFAULT_DAYS_AHEAD)
    check_out = check_in + timedelta(days=DEFAULT_NIGHTS)

    return SearchParams(
        chk_in=check_in.strftime("%Y-%m-%d"),
        chk_out=check_out.strftime("%Y-%m-%d"),
        rooms=DEFAULT_ROOMS,
        adults=DEFAULT_ADULTS,
        nights=DEFAULT_NIGHTS
    )


def get_multi_date_ranges() -> List[DateRange]:
    """
    Generate multiple date ranges to try for better price coverage.

    Returns a list of date ranges in priority order:
    1. +30 days (weekday)
    2. Next weekend (+35-42 days, Friday-Saturday)
    3. +60 days (weekday)
    4. +90 days (weekday)
    """
    today = datetime.now()
    date_ranges: List[DateRange] = []

    # 1. Primary: +30 days
    primary = today + timedelta(days=30)
    date_ranges.append(DateRange(
        label="+30d",
        chk_in=primary.strftime("%Y-%m-%d"),
        chk_out=(primary + timedelta(days=1)).strftime("%Y-%m-%d")
    ))

    # 2. Weekend: Find next Friday after +30 days
    days_until_friday = (4 - primary.weekday()) % 7
    if days_until_friday == 0:
        days_until_friday = 7  # Next Friday, not today
    weekend = primary + timedelta(days=days_until_friday)
    date_ranges.append(DateRange(
        label="weekend",
        chk_in=weekend.strftime("%Y-%m-%d"),
        chk_out=(weekend + timedelta(days=1)).strftime("%Y-%m-%d")
    ))

    # 3. +60 days
    d60 = today + timedelta(days=60)
    date_ranges.append(DateRange(
        label="+60d",
        chk_in=d60.strftime("%Y-%m-%d"),
        chk_out=(d60 + timedelta(days=1)).strftime("%Y-%m-%d")
    ))

    # 4. +90 days
    d90 = today + timedelta(days=90)
    date_ranges.append(DateRange(
        label="+90d",
        chk_in=d90.strftime("%Y-%m-%d"),
        chk_out=(d90 + timedelta(days=1)).strftime("%Y-%m-%d")
    ))

    return date_ranges


def try_multiple_dates(
    api: XoteloAPI,
    hotel_key: str,
    date_ranges: List[DateRange],
    rooms: int = 1,
    adults: int = 2
) -> Optional[Tuple[RateInfo, str]]:
    """
    Try multiple date ranges until a price is found.

    Args:
        api: XoteloAPI instance
        hotel_key: Hotel key to query
        date_ranges: List of date ranges to try
        rooms: Number of rooms
        adults: Adults per room

    Returns:
        Tuple of (RateInfo, date_label) if found, None otherwise
    """
    for date_range in date_ranges:
        rate_data = api.get_rates(
            hotel_key,
            date_range['chk_in'],
            date_range['chk_out'],
            rooms,
            adults
        )
        if rate_data:
            return (rate_data, date_range['label'])
        api.wait()

    return None


class HotelKeyData(TypedDict, total=False):
    """Type definition for hotel key data (new format)."""
    xotelo: str
    booking_url: str


def load_hotel_keys() -> Dict[str, Any]:
    """
    Load hotel name to key mappings from JSON database.

    Supports both old format (string) and new format (dict).
    Old: {"Hotel Name": "xotelo_key"}
    New: {"Hotel Name": {"xotelo": "xotelo_key", "booking_url": "..."}}
    """
    if not os.path.exists(HOTEL_KEYS_DB):
        logger.error("Hotel keys database not found: %s", HOTEL_KEYS_DB)
        return {}

    try:
        with open(HOTEL_KEYS_DB, 'r', encoding='utf-8') as f:
            keys: Dict[str, Any] = json.load(f)
        logger.info("Loaded %d hotel keys from %s", len(keys), HOTEL_KEYS_DB)
        return keys
    except (json.JSONDecodeError, OSError) as e:
        logger.error("Failed to load hotel keys: %s", e)
        return {}


def get_xotelo_key(hotel_data: Any) -> Optional[str]:
    """Extract Xotelo key from hotel data (supports old and new format)."""
    if isinstance(hotel_data, str):
        return hotel_data
    elif isinstance(hotel_data, dict):
        return hotel_data.get("xotelo")
    return None


def get_booking_url(hotel_data: Any) -> Optional[str]:
    """Extract Booking URL from hotel data."""
    if isinstance(hotel_data, dict):
        return hotel_data.get("booking_url")
    return None


def get_hotel_rates(
    hotel_key: str,
    chk_in: str,
    chk_out: str,
    rooms: int = 1,
    adults: int = 2,
    retries: int = 2
) -> Optional[Dict[str, Any]]:
    """
    Get hotel rates for specific dates and occupancy.

    This is a wrapper for backwards compatibility. New code should use
    XoteloAPI.get_rates() directly.
    """
    api = XoteloAPI(max_retries=retries)
    result = api.get_rates(hotel_key, chk_in, chk_out, rooms, adults)
    return dict(result) if result else None


def update_excel_with_prices(
    excel_hotels_with_prices: Dict[int, HotelPriceData],
    search_params: SearchParams,
    snapshot_date: str,
    multi_date: bool = False,
    cascade_mode: bool = False
) -> str:
    """Read Excel and update with prices, including snapshot date."""
    logger.info("Updating Excel file...")

    wb: Workbook = openpyxl.load_workbook(EXCEL_FILE)
    ws: Worksheet = wb.active

    # Find the last column with data and add new columns
    max_col = ws.max_column
    snapshot_col = max_col + 1
    price_col = max_col + 2
    provider_col = max_col + 3
    key_col = max_col + 4
    search_col = max_col + 5

    # Additional columns based on mode
    next_col = max_col + 6
    date_used_col = None
    source_col = None

    if multi_date:
        date_used_col = next_col
        next_col += 1

    if cascade_mode:
        source_col = next_col

    # Add headers with styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Use "Price_USD" for cascade mode (not just Xotelo)
    price_header = "Price_USD" if cascade_mode else "Xotelo_Price_USD"

    headers = [
        (snapshot_col, "Snapshot_Date"),
        (price_col, price_header),
        (provider_col, "Provider"),
        (key_col, "Hotel_Key"),
        (search_col, "Search_Params")
    ]

    if date_used_col:
        headers.append((date_used_col, "Date_Found"))

    if source_col:
        headers.append((source_col, "Source"))

    for col, header_text in headers:
        cell = ws.cell(row=1, column=col, value=header_text)
        cell.fill = header_fill
        cell.font = header_font

    # Create search info string
    search_info = f"{search_params['chk_in']} to {search_params['chk_out']} | {search_params['rooms']}rm/{search_params['adults']}ad"
    if multi_date:
        search_info += " (multi-date)"
    if cascade_mode:
        search_info += " (cascade)"

    # Update rows with matched data
    for row_num, hotel_data in excel_hotels_with_prices.items():
        ws.cell(row=row_num, column=snapshot_col, value=snapshot_date)
        ws.cell(row=row_num, column=price_col, value=hotel_data.get('price'))
        ws.cell(row=row_num, column=provider_col, value=hotel_data.get('provider'))
        ws.cell(row=row_num, column=key_col, value=hotel_data.get('hotel_key'))
        ws.cell(row=row_num, column=search_col, value=search_info)
        if date_used_col:
            ws.cell(row=row_num, column=date_used_col, value=hotel_data.get('date_used', ''))
        if source_col:
            ws.cell(row=row_num, column=source_col, value=hotel_data.get('source', ''))

    # Generate output filename with date
    output_file = f"PRTC_Hotels_Prices_{snapshot_date}.xlsx"
    wb.save(output_file)
    logger.info("Saved to: %s", output_file)

    return output_file


def main() -> None:
    """Main execution function."""
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Hotel Price Updater - Xotelo API')
    parser.add_argument('--auto', action='store_true',
                        help='Run in automatic mode with default parameters (no prompts)')
    parser.add_argument('--multi-date', action='store_true',
                        help='Try multiple date ranges to maximize price coverage')
    parser.add_argument('--cascade', action='store_true',
                        help='Use cascade pipeline (Xotelo -> SerpApi -> Apify) for max coverage')
    parser.add_argument('--limit', type=int, default=0,
                        help='Limit number of hotels to process (for testing)')
    args = parser.parse_args()

    # Get snapshot date (when the data was collected)
    snapshot_date = datetime.now().strftime("%Y-%m-%d")

    print("=" * 70)
    print("Hotel Price Updater - Xotelo API (Key-Based)")
    print("=" * 70)
    print(f"Snapshot Date: {snapshot_date}")

    if args.auto:
        print("Mode: AUTOMATIC (using default parameters)")
        search_params = get_auto_params()
        print(f"   Check-in:  {search_params['chk_in']} (+{DEFAULT_DAYS_AHEAD} days)")
        print(f"   Check-out: {search_params['chk_out']}")
        print(f"   Rooms: {search_params['rooms']}, Adults/room: {search_params['adults']}")
    else:
        print("Mode: INTERACTIVE")
        search_params_result = get_user_input()
        if not search_params_result:
            return
        search_params = search_params_result

    # Multi-date mode setup
    date_ranges: Optional[List[DateRange]] = None
    if args.multi_date:
        print("\nMulti-date mode: ENABLED")
        date_ranges = get_multi_date_ranges()
        print("   Date ranges to try:")
        for dr in date_ranges:
            print(f"     - {dr['label']}: {dr['chk_in']} to {dr['chk_out']}")

    # Cascade mode setup
    cascade_provider: Optional[CascadePriceProvider] = None
    if args.cascade:
        if not CASCADE_AVAILABLE:
            logger.error("Cascade mode requires price_providers package. Install dependencies:")
            logger.error("  pip install google-search-results apify-client python-dotenv")
            return

        print("\nCascade mode: ENABLED")

        # Initialize providers
        providers = []

        # 1. Xotelo (primary, always available)
        xotelo = XoteloProvider()
        if date_ranges:
            xotelo.set_multi_date_ranges(date_ranges)
        providers.append(xotelo)
        print("   [1] Xotelo: OK (primary)")

        # 2. SerpApi (Google Hotels)
        serpapi = SerpApiProvider()
        if serpapi.is_available():
            providers.append(serpapi)
            print("   [2] SerpApi: OK (Google Hotels)")
        else:
            print("   [2] SerpApi: SKIP (no SERPAPI_KEY)")

        # 3. Apify (Booking.com)
        apify = ApifyProvider()
        if apify.is_available():
            providers.append(apify)
            print("   [3] Apify: OK (Booking.com)")
        else:
            print("   [3] Apify: SKIP (no APIFY_TOKEN)")

        # Initialize cache
        cache = PriceCache(
            cache_file=config.CACHE_FILE,
            ttl_hours=config.CACHE_TTL_HOURS
        )
        cache_stats = cache.get_stats()
        print(f"   Cache: {cache_stats['valid_entries']} valid entries")

        # Create cascade provider
        cascade_provider = CascadePriceProvider(providers, cache)

    print("=" * 70)

    # Step 1: Load hotel keys from database
    print("\n[STEP 1] Loading hotel keys database...")
    hotel_keys = load_hotel_keys()

    if not hotel_keys:
        logger.error("No hotel keys loaded. Exiting.")
        return

    # Step 2: Load Excel and process hotels
    print("\n[STEP 2] Loading Excel and fetching prices by key...")
    wb: Workbook = openpyxl.load_workbook(EXCEL_FILE)
    ws: Worksheet = wb.active

    # Initialize API client
    api = get_client()

    # Initialize counters
    hotels_total = 0
    hotels_with_key = 0
    hotels_with_prices = 0
    excel_hotels_with_prices: Dict[int, HotelPriceData] = {}

    # Track date usage stats for multi-date mode
    date_stats: Dict[str, int] = {}

    for row in range(2, ws.max_row + 1):
        hotel_name = ws.cell(row=row, column=1).value

        if not hotel_name:
            continue

        hotel_name = str(hotel_name).strip()
        hotels_total += 1

        # Check limit for testing
        if args.limit > 0 and hotels_total > args.limit:
            print(f"\n[LIMIT] Reached limit of {args.limit} hotels")
            break

        hotel_data = hotel_keys.get(hotel_name)
        hotel_key = get_xotelo_key(hotel_data)
        booking_url = get_booking_url(hotel_data)

        # CASCADE MODE
        if args.cascade and cascade_provider:
            hotels_with_key += 1  # In cascade mode, we try all hotels
            print(f"\n[{hotels_total}] {hotel_name}")
            if hotel_key:
                print(f"    Key: {hotel_key}")
            else:
                print("    Key: None (will try SerpApi/Apify)")
            if booking_url:
                print(f"    Booking URL: {booking_url[:50]}...")

            result = cascade_provider.get_price(
                hotel_name,
                hotel_key,
                search_params['chk_in'],
                search_params['chk_out'],
                search_params['rooms'],
                search_params['adults'],
                booking_url=booking_url
            )

            if result:
                price = result['price']
                provider = result['provider']
                source = result['source']
                cached = result['cached']
                cache_label = " [cached]" if cached else ""

                print(f"    Price: ${price:.2f} ({provider}) via {source}{cache_label}")

                excel_hotels_with_prices[row] = HotelPriceData(
                    price=price,
                    provider=provider,
                    hotel_key=hotel_key or '',
                    source=source
                )
                hotels_with_prices += 1
            else:
                print("    No price available (all providers failed)")
                excel_hotels_with_prices[row] = HotelPriceData(
                    price=None,
                    provider='N/A',
                    hotel_key=hotel_key or '',
                    source='none'
                )

        # NON-CASCADE MODE (original behavior)
        elif hotel_key:  # hotel_key is already extracted above
            hotels_with_key += 1
            print(f"\n[{hotels_with_key}] {hotel_name}")
            print(f"    Key: {hotel_key}")
            print("    Fetching price...")

            if args.multi_date and date_ranges:
                # Try multiple dates
                result = try_multiple_dates(
                    api,
                    hotel_key,
                    date_ranges,
                    search_params['rooms'],
                    search_params['adults']
                )

                if result:
                    rate_data, date_label = result
                    price = rate_data['rate']
                    provider = rate_data['provider']
                    print(f"    Price: ${price:.2f} ({provider}) [found: {date_label}]")

                    excel_hotels_with_prices[row] = HotelPriceData(
                        price=price,
                        provider=provider,
                        hotel_key=hotel_key,
                        date_used=date_label
                    )
                    hotels_with_prices += 1
                    date_stats[date_label] = date_stats.get(date_label, 0) + 1
                else:
                    print("    No price available (tried all dates)")
                    excel_hotels_with_prices[row] = HotelPriceData(
                        price=None,
                        provider='N/A',
                        hotel_key=hotel_key,
                        date_used=''
                    )
            else:
                # Single date mode (original behavior)
                rate_data = api.get_rates(
                    hotel_key,
                    search_params['chk_in'],
                    search_params['chk_out'],
                    search_params['rooms'],
                    search_params['adults']
                )

                if rate_data:
                    price = rate_data['rate']
                    provider = rate_data['provider']
                    print(f"    Price: ${price:.2f} ({provider})")

                    excel_hotels_with_prices[row] = HotelPriceData(
                        price=price,
                        provider=provider,
                        hotel_key=hotel_key
                    )
                    hotels_with_prices += 1
                else:
                    print("    No price available")
                    excel_hotels_with_prices[row] = HotelPriceData(
                        price=None,
                        provider='N/A',
                        hotel_key=hotel_key
                    )
            api.wait()
        else:
            print(f"\n[SKIP] {hotel_name} - No key in database")

    print(f"\n[STATS] Hotels in Excel: {hotels_total}")
    print(f"[STATS] Hotels with keys: {hotels_with_key}")
    print(f"[STATS] Hotels with prices: {hotels_with_prices}")

    # Show cascade statistics if in cascade mode
    if args.cascade and cascade_provider:
        print("\n" + cascade_provider.get_stats_summary())
    elif args.multi_date and date_stats:
        print("\n[STATS] Prices found by date range:")
        for date_label, count in sorted(date_stats.items(), key=lambda x: -x[1]):
            print(f"   {date_label}: {count} hotels")

    # Step 3: Update Excel
    if excel_hotels_with_prices:
        output_file = update_excel_with_prices(
            excel_hotels_with_prices,
            search_params,
            snapshot_date,
            multi_date=args.multi_date,
            cascade_mode=args.cascade
        )
        print("\n" + "=" * 70)
        print("[COMPLETE]")
        print(f"Output: {output_file}")
        print(f"{hotels_with_prices} hotels with prices")
        print(f"{hotels_with_key - hotels_with_prices} hotels without prices")
        print(f"{hotels_total - hotels_with_key} hotels without keys")
        print("=" * 70)
    else:
        logger.warning("No hotels matched or priced")


if __name__ == "__main__":
    main()
