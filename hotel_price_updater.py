"""
Hotel Price Updater - Google Hotel API2 via RapidAPI
Fetches hotel prices and updates the PRTC Endorsed Hotels Excel file.
With rate limiting handling for free tier.
"""

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill
from difflib import SequenceMatcher
import time

# RapidAPI Configuration - Google Hotel API2
RAPIDAPI_KEY = "61a6806b0bmsh8ba9a803cf323cbp18dabbjsn2baf2c831718"
RAPIDAPI_HOST = "google-hotel-api2.p.rapidapi.com"
BASE_URL = "https://google-hotel-api2.p.rapidapi.com"

# File paths
EXCEL_FILE = "PRTC Endorsed Hotels (12.25).xlsx"
OUTPUT_FILE = "PRTC Endorsed Hotels (12.25) - Updated Prices.xlsx"

# Headers for RapidAPI
HEADERS = {
    "x-rapidapi-key": RAPIDAPI_KEY,
    "x-rapidapi-host": RAPIDAPI_HOST
}

# Request timeout and rate limiting
TIMEOUT = 30
REQUEST_DELAY = 3  # Seconds between API calls (free tier rate limit)


def search_city(keyword, retries=2):
    """Search for a city and get its code."""
    print(f"[CITY] Searching: {keyword}")
    
    url = f"{BASE_URL}/hotel/reference-data-locations-cities"
    params = {"keyword": keyword}
    
    for attempt in range(retries):
        try:
            response = requests.get(url, headers=HEADERS, params=params, timeout=TIMEOUT)
            
            if response.status_code == 429:
                print(f"  [RATE LIMIT] Waiting 10 seconds...")
                time.sleep(10)
                continue
                
            response.raise_for_status()
            data = response.json()
            
            if data and 'data' in data:
                cities = data['data']
                # Look for Puerto Rico specifically
                for city in cities:
                    address = city.get('address', {})
                    if address.get('countryCode') == 'PR':
                        city_code = city.get('iataCode')
                        city_name = city.get('name', keyword)
                        print(f"  [OK] Found: {city_name} (Code: {city_code})")
                        return city_code, city_name
                
                # If no PR result, return first result
                if cities:
                    city = cities[0]
                    city_code = city.get('iataCode')
                    city_name = city.get('name', keyword)
                    print(f"  [OK] Found: {city_name} (Code: {city_code})")
                    return city_code, city_name
            
            print(f"  [WARN] No city found for: {keyword}")
            return None, None
            
        except requests.exceptions.Timeout:
            print(f"  [TIMEOUT] Request timed out for: {keyword}")
            return None, None
        except requests.exceptions.RequestException as e:
            if attempt < retries - 1:
                print(f"  [RETRY] Attempt {attempt + 1} failed, retrying...")
                time.sleep(5)
            else:
                print(f"  [ERROR] Search failed: {e}")
                return None, None
    
    return None, None


def get_hotels_by_city(city_code, retries=2):
    """Get list of hotels in a city."""
    print(f"[HOTELS] Fetching hotels for city: {city_code}")
    
    url = f"{BASE_URL}/hotel/reference-data-locations-hotels-by-city"
    params = {"cityCode": city_code}
    
    for attempt in range(retries):
        try:
            response = requests.get(url, headers=HEADERS, params=params, timeout=TIMEOUT)
            
            if response.status_code == 429:
                print(f"  [RATE LIMIT] Waiting 10 seconds...")
                time.sleep(10)
                continue
                
            response.raise_for_status()
            data = response.json()
            
            hotels = []
            hotel_data = data.get('data', data) if isinstance(data, dict) else data
            
            if hotel_data and isinstance(hotel_data, list):
                for hotel in hotel_data:
                    hotels.append({
                        'id': hotel.get('hotelId'),
                        'name': hotel.get('name', 'Unknown'),
                        'city': city_code
                    })
                print(f"  [OK] Found {len(hotels)} hotels")
            
            return hotels
            
        except requests.exceptions.Timeout:
            print(f"  [TIMEOUT] Request timed out for city: {city_code}")
            return []
        except requests.exceptions.RequestException as e:
            if attempt < retries - 1:
                print(f"  [RETRY] Attempt {attempt + 1} failed, retrying...")
                time.sleep(5)
            else:
                print(f"  [ERROR] Hotel list failed: {e}")
                return []
    
    return []


def get_hotel_offers(hotel_ids, check_in="2026-01-30", check_out="2026-01-31", retries=2):
    """Get hotel offers/prices for specific hotels."""
    ids_list = hotel_ids if isinstance(hotel_ids, list) else [hotel_ids]
    print(f"[PRICES] Fetching prices for {len(ids_list)} hotel(s)...")
    
    url = f"{BASE_URL}/hotel/shopping-hotel-offers"
    ids_str = ",".join(ids_list)
    
    params = {
        "hotelIds": ids_str,
        "checkInDate": check_in,
        "checkOutDate": check_out,
        "adults": "2",
        "roomQuantity": "1",
        "currency": "USD",
        "view": "FULL"
    }
    
    for attempt in range(retries):
        try:
            response = requests.get(url, headers=HEADERS, params=params, timeout=TIMEOUT)
            
            if response.status_code == 429:
                print(f"  [RATE LIMIT] Waiting 10 seconds...")
                time.sleep(10)
                continue
                
            response.raise_for_status()
            data = response.json()
            
            results = []
            offers_data = data.get('data', data) if isinstance(data, dict) else data
            
            if offers_data and isinstance(offers_data, list):
                for offer in offers_data:
                    hotel_info = offer.get('hotel', {})
                    hotel_name = hotel_info.get('name', 'Unknown')
                    hotel_id = hotel_info.get('hotelId', '')
                    
                    # Get the lowest price from all offers
                    hotel_offers = offer.get('offers', [])
                    lowest_price = None
                    
                    for o in hotel_offers:
                        price_info = o.get('price', {})
                        total = price_info.get('total')
                        if total:
                            try:
                                price_val = float(total)
                                if lowest_price is None or price_val < lowest_price:
                                    lowest_price = price_val
                            except (ValueError, TypeError):
                                pass
                    
                    if lowest_price:
                        results.append({
                            'id': hotel_id,
                            'name': hotel_name,
                            'price': lowest_price,
                            'currency': 'USD'
                        })
                        print(f"  [PRICE] {hotel_name}: ${lowest_price}")
            
            return results
            
        except requests.exceptions.Timeout:
            print(f"  [TIMEOUT] Request timed out")
            return []
        except requests.exceptions.RequestException as e:
            if attempt < retries - 1:
                print(f"  [RETRY] Attempt {attempt + 1} failed, retrying...")
                time.sleep(5)
            else:
                print(f"  [ERROR] Offers failed: {e}")
                return []
    
    return []


def similarity_ratio(a, b):
    """Calculate similarity between two strings."""
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def find_matching_hotel(hotel_name, api_hotels):
    """Find the best matching hotel from API results."""
    best_match = None
    best_score = 0
    
    for hotel in api_hotels:
        score = similarity_ratio(hotel_name, hotel['name'])
        if score > best_score and score > 0.4:
            best_score = score
            best_match = hotel
    
    return best_match, best_score


def update_excel_with_prices(all_hotels):
    """Read Excel, match hotels, and update with prices."""
    print("\n[EXCEL] Updating Excel file...")
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Find the last column with data and add new columns
    max_col = ws.max_column
    price_col = max_col + 1
    match_col = max_col + 2
    
    # Add headers with styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    ws.cell(row=1, column=price_col, value="API_Price_USD").fill = header_fill
    ws.cell(row=1, column=price_col).font = header_font
    ws.cell(row=1, column=match_col, value="API_Match_Name").fill = header_fill
    ws.cell(row=1, column=match_col).font = header_font
    
    # Process each hotel
    hotels_matched = 0
    hotels_total = 0
    
    for row in range(2, ws.max_row + 1):
        hotel_name = ws.cell(row=row, column=1).value
        
        if not hotel_name:
            continue
        
        hotels_total += 1
        
        # Try to find matching hotel in API results
        match, score = find_matching_hotel(hotel_name, all_hotels)
        
        if match:
            ws.cell(row=row, column=price_col, value=match['price'])
            ws.cell(row=row, column=match_col, value=match['name'])
            hotels_matched += 1
            print(f"  [MATCH] {hotel_name} -> {match['name']} (${match['price']}, {score:.0%})")
    
    # Save the updated workbook
    wb.save(OUTPUT_FILE)
    print(f"\n[SAVE] Saved to: {OUTPUT_FILE}")
    print(f"[STATS] Matched {hotels_matched}/{hotels_total} hotels with prices")
    
    return hotels_matched, hotels_total


def main():
    """Main execution function."""
    print("=" * 60)
    print("Hotel Price Updater - Google Hotel API2")
    print("=" * 60)
    
    # Step 1: Search for San Juan, Puerto Rico
    print("\n[STEP 1] Finding San Juan, Puerto Rico...")
    
    city_code, city_name = search_city("San Juan")
    time.sleep(REQUEST_DELAY)
    
    if not city_code:
        print("[ERROR] Could not find San Juan. Exiting.")
        return
    
    # Step 2: Get hotels in San Juan
    print(f"\n[STEP 2] Getting hotels in {city_name}...")
    hotels = get_hotels_by_city(city_code)
    time.sleep(REQUEST_DELAY)
    
    if not hotels:
        print("[ERROR] No hotels found. Exiting.")
        return
    
    print(f"[INFO] Found {len(hotels)} hotels in {city_name}")
    
    # Step 3: Get prices for hotels (in batches to avoid rate limits)
    print("\n[STEP 3] Fetching hotel prices...")
    all_prices = []
    
    hotel_ids = [h['id'] for h in hotels if h.get('id')]
    batch_size = 5  # Small batches to avoid rate limits
    
    for i in range(0, min(len(hotel_ids), 20), batch_size):  # Limit to first 20 hotels
        batch = hotel_ids[i:i+batch_size]
        print(f"\n[BATCH] Processing hotels {i+1} to {i+len(batch)}...")
        
        prices = get_hotel_offers(batch)
        all_prices.extend(prices)
        
        print(f"[WAIT] Waiting {REQUEST_DELAY} seconds for rate limit...")
        time.sleep(REQUEST_DELAY)
    
    print(f"\n[INFO] Got prices for {len(all_prices)} hotels")
    
    # Step 4: Update Excel
    if all_prices:
        matched, total = update_excel_with_prices(all_prices)
        print("\n" + "=" * 60)
        print("[COMPLETE]")
        print(f"[RESULTS] {matched}/{total} hotels matched with prices")
        print("=" * 60)
    else:
        print("\n[WARN] No hotel prices retrieved from API")


if __name__ == "__main__":
    main()
