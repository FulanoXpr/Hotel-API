"""
Unit tests for extract_all_hotels.py
"""
import pytest
import json
import os
import sys
from unittest.mock import patch, MagicMock

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import extract_all_hotels as extractor
from xotelo_api import XoteloAPI


class TestExtractHotelData:
    """Tests for extract_hotel_data function."""

    def test_extracts_all_fields(self):
        raw_hotels = [
            {
                'name': 'Test Hotel',
                'key': 'test-key',
                'url': 'https://example.com/hotel',
                'accommodation_type': 'hotel',
                'review_summary': {
                    'rating': 4.5,
                    'count': 100
                }
            }
        ]

        result = extractor.extract_hotel_data(raw_hotels)

        assert len(result) == 1
        assert result[0]['name'] == 'Test Hotel'
        assert result[0]['key'] == 'test-key'
        assert result[0]['url'] == 'https://example.com/hotel'
        assert result[0]['accommodation_type'] == 'hotel'
        assert result[0]['rating'] == 4.5
        assert result[0]['review_count'] == 100

    def test_handles_missing_review_summary(self):
        raw_hotels = [
            {
                'name': 'Hotel Without Reviews',
                'key': 'no-review-key',
                'url': 'https://example.com',
                'accommodation_type': 'inn'
            }
        ]

        result = extractor.extract_hotel_data(raw_hotels)

        assert len(result) == 1
        assert result[0]['rating'] is None
        assert result[0]['review_count'] is None

    def test_handles_empty_list(self):
        result = extractor.extract_hotel_data([])

        assert result == []

    def test_handles_multiple_hotels(self):
        raw_hotels = [
            {'name': 'Hotel A', 'key': 'a', 'url': '', 'accommodation_type': ''},
            {'name': 'Hotel B', 'key': 'b', 'url': '', 'accommodation_type': ''},
            {'name': 'Hotel C', 'key': 'c', 'url': '', 'accommodation_type': ''}
        ]

        result = extractor.extract_hotel_data(raw_hotels)

        assert len(result) == 3
        assert result[0]['name'] == 'Hotel A'
        assert result[2]['name'] == 'Hotel C'


class TestSaveToJson:
    """Tests for save_to_json function."""

    def test_saves_valid_json(self, tmp_path):
        hotel_data = [
            extractor.HotelData(
                name='Test Hotel',
                key='test-key',
                url='https://example.com',
                accommodation_type='hotel',
                rating=4.0,
                review_count=50
            )
        ]
        filename = str(tmp_path / "test_output.json")

        extractor.save_to_json(hotel_data, filename)

        with open(filename, 'r', encoding='utf-8') as f:
            loaded = json.load(f)

        assert len(loaded) == 1
        assert loaded[0]['name'] == 'Test Hotel'

    def test_handles_unicode_characters(self, tmp_path):
        hotel_data = [
            extractor.HotelData(
                name='Hotel Niño',
                key='key',
                url='',
                accommodation_type='',
                rating=None,
                review_count=None
            )
        ]
        filename = str(tmp_path / "unicode_test.json")

        extractor.save_to_json(hotel_data, filename)

        with open(filename, 'r', encoding='utf-8') as f:
            loaded = json.load(f)

        assert loaded[0]['name'] == 'Hotel Niño'


class TestSaveToExcel:
    """Tests for save_to_excel function."""

    def test_creates_excel_file(self, tmp_path):
        hotel_data = [
            extractor.HotelData(
                name='Test Hotel',
                key='test-key',
                url='https://example.com',
                accommodation_type='hotel',
                rating=4.0,
                review_count=50
            )
        ]
        filename = str(tmp_path / "test_output.xlsx")

        extractor.save_to_excel(hotel_data, filename)

        assert os.path.exists(filename)

    def test_excel_has_correct_headers(self, tmp_path):
        import openpyxl

        hotel_data = [
            extractor.HotelData(
                name='Test',
                key='key',
                url='url',
                accommodation_type='type',
                rating=1.0,
                review_count=1
            )
        ]
        filename = str(tmp_path / "headers_test.xlsx")

        extractor.save_to_excel(hotel_data, filename)

        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers == ["#", "Hotel Name", "Key", "URL", "Type", "Rating", "Reviews"]

    def test_excel_has_correct_data(self, tmp_path):
        import openpyxl

        hotel_data = [
            extractor.HotelData(
                name='Sample Hotel',
                key='sample-key',
                url='https://sample.com',
                accommodation_type='resort',
                rating=4.5,
                review_count=200
            )
        ]
        filename = str(tmp_path / "data_test.xlsx")

        extractor.save_to_excel(hotel_data, filename)

        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # Check data in row 2
        assert ws.cell(row=2, column=1).value == 1  # Index
        assert ws.cell(row=2, column=2).value == 'Sample Hotel'
        assert ws.cell(row=2, column=3).value == 'sample-key'
        assert ws.cell(row=2, column=4).value == 'https://sample.com'
        assert ws.cell(row=2, column=5).value == 'resort'
        assert ws.cell(row=2, column=6).value == 4.5
        assert ws.cell(row=2, column=7).value == 200


class TestMainIntegration:
    """Integration tests for main function."""

    @patch.object(XoteloAPI, 'list_hotels')
    @patch.object(XoteloAPI, 'wait')
    @patch('extract_all_hotels.save_to_json')
    @patch('extract_all_hotels.save_to_excel')
    def test_fetches_all_hotels_with_pagination(
        self, mock_save_excel, mock_save_json, mock_wait, mock_list
    ):
        # Simulate pagination: first call returns 100 hotels, second returns remaining
        mock_list.side_effect = [
            ([{'name': f'Hotel {i}', 'key': f'key-{i}', 'url': '', 'accommodation_type': ''}
              for i in range(100)], 150),
            ([{'name': f'Hotel {i}', 'key': f'key-{i}', 'url': '', 'accommodation_type': ''}
              for i in range(100, 150)], 150),
        ]

        extractor.main()

        # Should have called list_hotels twice for pagination
        assert mock_list.call_count == 2
        # Should have saved data
        assert mock_save_json.called
        assert mock_save_excel.called

    @patch.object(XoteloAPI, 'list_hotels')
    def test_handles_empty_response(self, mock_list):
        mock_list.return_value = ([], 0)

        # Should not raise, just return
        extractor.main()

        assert mock_list.called


class TestConstants:
    """Tests for module constants."""

    def test_output_json_has_extension(self):
        assert extractor.OUTPUT_JSON.endswith('.json')

    def test_output_excel_has_extension(self):
        assert extractor.OUTPUT_EXCEL.endswith('.xlsx')
