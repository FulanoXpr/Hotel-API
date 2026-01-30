"""
Integration tests for Excel file handling.
"""
import pytest
import json
import os
import sys
from datetime import datetime
from unittest.mock import patch, MagicMock
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import xotelo_price_updater as updater


@pytest.fixture
def sample_excel(tmp_path):
    """Create a sample Excel file for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add header
    ws.cell(row=1, column=1, value="Hotel Name")
    ws.cell(row=1, column=2, value="Location")

    # Add sample hotels
    hotels = [
        ("Hotel Alpha", "San Juan"),
        ("Hotel Beta", "Ponce"),
        ("Hotel Gamma", "Mayaguez"),
    ]

    for i, (name, location) in enumerate(hotels, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=location)

    filepath = tmp_path / "test_hotels.xlsx"
    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def sample_keys_db(tmp_path):
    """Create a sample hotel keys database."""
    keys = {
        "Hotel Alpha": "g147319-d111111",
        "Hotel Beta": "g147319-d222222"
        # Note: Hotel Gamma intentionally not mapped
    }

    filepath = tmp_path / "test_keys.json"
    filepath.write_text(json.dumps(keys))
    return str(filepath)


class TestExcelReading:
    """Tests for reading Excel files."""

    def test_reads_hotel_names_from_excel(self, sample_excel):
        wb = openpyxl.load_workbook(sample_excel)
        ws = wb.active

        hotels = []
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if name:
                hotels.append(name.strip())

        assert len(hotels) == 3
        assert "Hotel Alpha" in hotels
        assert "Hotel Beta" in hotels
        assert "Hotel Gamma" in hotels


class TestUpdateExcelWithPrices:
    """Tests for update_excel_with_prices function."""

    def test_adds_price_columns(self, sample_excel):
        # Setup
        with patch.object(updater, 'EXCEL_FILE', sample_excel):
            excel_hotels_with_prices = {
                2: {'price': 150.00, 'provider': 'Booking.com', 'hotel_key': 'g147319-d111111'},
                3: {'price': 120.50, 'provider': 'Agoda', 'hotel_key': 'g147319-d222222'},
            }

            search_params = {
                'chk_in': '2026-03-01',
                'chk_out': '2026-03-02',
                'rooms': 1,
                'adults': 2
            }

            snapshot_date = '2026-01-30'

            # Execute
            output_file = updater.update_excel_with_prices(
                excel_hotels_with_prices,
                search_params,
                snapshot_date
            )

        # Verify
        assert os.path.exists(output_file)

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        # Check that new columns were added (original had 2 columns)
        assert ws.max_column >= 5

        # Find the price column
        price_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "Xotelo_Price_USD":
                price_col = col
                break

        assert price_col is not None

        # Check prices were written
        assert ws.cell(row=2, column=price_col).value == 150.00
        assert ws.cell(row=3, column=price_col).value == 120.50

        # Cleanup
        os.remove(output_file)

    def test_includes_snapshot_date(self, sample_excel):
        with patch.object(updater, 'EXCEL_FILE', sample_excel):
            excel_hotels_with_prices = {
                2: {'price': 100, 'provider': 'Test', 'hotel_key': 'key-1'},
            }

            search_params = {
                'chk_in': '2026-03-01',
                'chk_out': '2026-03-02',
                'rooms': 1,
                'adults': 2
            }

            snapshot_date = '2026-01-30'

            output_file = updater.update_excel_with_prices(
                excel_hotels_with_prices,
                search_params,
                snapshot_date
            )

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        # Find snapshot column
        snapshot_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "Snapshot_Date":
                snapshot_col = col
                break

        assert snapshot_col is not None
        assert ws.cell(row=2, column=snapshot_col).value == '2026-01-30'

        # Cleanup
        os.remove(output_file)

    def test_output_filename_includes_date(self, sample_excel):
        with patch.object(updater, 'EXCEL_FILE', sample_excel):
            excel_hotels_with_prices = {
                2: {'price': 100, 'provider': 'Test', 'hotel_key': 'key-1'},
            }

            search_params = {
                'chk_in': '2026-03-01',
                'chk_out': '2026-03-02',
                'rooms': 1,
                'adults': 2
            }

            snapshot_date = '2026-01-30'

            output_file = updater.update_excel_with_prices(
                excel_hotels_with_prices,
                search_params,
                snapshot_date
            )

        assert 'PRTC_Hotels_Prices_2026-01-30.xlsx' in output_file

        # Cleanup
        os.remove(output_file)


class TestKeyManagerExcelReading:
    """Tests for key_manager Excel reading."""

    def test_get_excel_hotels_returns_sorted_unique_list(self, sample_excel):
        import key_manager

        with patch.object(key_manager, 'EXCEL_FILE', sample_excel):
            hotels = key_manager.get_excel_hotels()

        assert len(hotels) == 3
        # Should be sorted
        assert hotels == sorted(hotels)

    def test_handles_missing_file(self, tmp_path):
        import key_manager

        with patch.object(key_manager, 'EXCEL_FILE', str(tmp_path / "nonexistent.xlsx")):
            hotels = key_manager.get_excel_hotels()

        assert hotels == []
