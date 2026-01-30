"""
Extract All Hotels from Puerto Rico - Xotelo API
Extracts all hotels with their Key and TripAdvisor URL.
"""
from __future__ import annotations

import json
import logging
import sys
from typing import Any, Dict, List, Optional, TypedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import config
from xotelo_api import XoteloAPI

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(levelname)s] %(message)s'
)
logger = logging.getLogger(__name__)

# Force immediate output
sys.stdout.reconfigure(line_buffering=True)

# Output files
OUTPUT_JSON = "all_hotels_puerto_rico.json"
OUTPUT_EXCEL = "all_hotels_puerto_rico.xlsx"


class HotelData(TypedDict, total=False):
    """Type definition for extracted hotel data."""
    name: str
    key: str
    url: str
    accommodation_type: str
    rating: Optional[float]
    review_count: Optional[int]


def extract_hotel_data(hotels: List[Dict[str, Any]]) -> List[HotelData]:
    """
    Extract relevant data from raw hotel list.

    Args:
        hotels: List of raw hotel dictionaries from API

    Returns:
        List of HotelData with extracted fields
    """
    hotel_data: List[HotelData] = []
    for hotel in hotels:
        review_summary = hotel.get("review_summary", {})
        hotel_data.append(HotelData(
            name=hotel.get("name", ""),
            key=hotel.get("key", ""),
            url=hotel.get("url", ""),
            accommodation_type=hotel.get("accommodation_type", ""),
            rating=review_summary.get("rating") if review_summary else None,
            review_count=review_summary.get("count") if review_summary else None
        ))
    return hotel_data


def save_to_json(hotel_data: List[HotelData], filename: str) -> None:
    """
    Save hotel data to JSON file.

    Args:
        hotel_data: List of hotel data to save
        filename: Output filename
    """
    logger.info("Saving to %s...", filename)
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(hotel_data, f, ensure_ascii=False, indent=2)
    logger.info("Saved %d hotels to JSON", len(hotel_data))


def save_to_excel(hotel_data: List[HotelData], filename: str) -> None:
    """
    Save hotel data to Excel file.

    Args:
        hotel_data: List of hotel data to save
        filename: Output filename
    """
    logger.info("Saving to %s...", filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hotels Puerto Rico"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    headers = ["#", "Hotel Name", "Key", "URL", "Type", "Rating", "Reviews"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for idx, hotel in enumerate(hotel_data, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=hotel["name"])
        ws.cell(row=idx + 1, column=3, value=hotel["key"])
        ws.cell(row=idx + 1, column=4, value=hotel["url"])
        ws.cell(row=idx + 1, column=5, value=hotel["accommodation_type"])
        ws.cell(row=idx + 1, column=6, value=hotel["rating"])
        ws.cell(row=idx + 1, column=7, value=hotel["review_count"])

    # Adjust column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10

    wb.save(filename)
    logger.info("Saved %d hotels to Excel", len(hotel_data))


def main() -> None:
    """Main execution function."""
    print("=" * 70)
    print("Extract All Hotels from Puerto Rico - Xotelo API")
    print("=" * 70)
    print("Location: Puerto Rico (entire island)")
    print("=" * 70)

    # Initialize API client with 1 second delay
    api = XoteloAPI(delay=1.0)

    # Get all hotels from Puerto Rico (paginated)
    print("\n[STEP 1] Fetching ALL hotels from Puerto Rico...")
    all_hotels: List[Dict[str, Any]] = []
    offset = 0
    limit = 100
    total_count = 0

    while True:
        print(f"  [FETCH] Getting hotels {offset + 1} to {offset + limit}...")
        hotels, total = api.list_hotels(
            location_key=config.LOCATION_KEY,
            limit=limit,
            offset=offset
        )

        if total > 0:
            total_count = total

        if not hotels:
            if offset == 0:
                logger.error("No hotels found from API. Exiting.")
                return
            break

        all_hotels.extend(hotels)

        print(f"  [PROGRESS] Collected {len(all_hotels)}/{total_count} hotels")

        # Check if we've collected all hotels
        if offset + limit >= total_count:
            print(f"  [DONE] All {total_count} hotels collected!")
            break

        offset += limit
        api.wait()

    print(f"\n[INFO] Total hotels collected: {len(all_hotels)}")

    # Extract relevant data
    print("\n[STEP 2] Extracting Key and URL data...")
    hotel_data = extract_hotel_data(all_hotels)

    # Save to JSON
    print(f"\n[STEP 3] Saving to {OUTPUT_JSON}...")
    save_to_json(hotel_data, OUTPUT_JSON)

    # Save to Excel
    print(f"\n[STEP 4] Saving to {OUTPUT_EXCEL}...")
    save_to_excel(hotel_data, OUTPUT_EXCEL)

    # Summary
    print("\n" + "=" * 70)
    print("[COMPLETE] Extraction finished!")
    print(f"  - Total hotels: {len(hotel_data)}")
    print(f"  - JSON file: {OUTPUT_JSON}")
    print(f"  - Excel file: {OUTPUT_EXCEL}")
    print("=" * 70)

    # Show first 5 hotels as sample
    print("\n[SAMPLE] First 5 hotels:")
    for i, hotel in enumerate(hotel_data[:5], 1):
        print(f"  {i}. {hotel['name']}")
        print(f"     Key: {hotel['key']}")
        url = hotel['url']
        print(f"     URL: {url[:60]}..." if len(url) > 60 else f"     URL: {url}")


if __name__ == "__main__":
    main()
